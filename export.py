"""
KOL 스코어카드 - Excel 내보내기 모듈
openpyxl 기반으로 3시트 Excel 생성
"""
from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─── 색상 팔레트 ─────────────────────────────────────────────
C_NAVY    = "1A3A5C"
C_BLUE    = "2E75B6"
C_TEAL    = "1F7A8C"
C_LGRAY   = "F2F2F2"
C_MGRAY   = "D9D9D9"
C_MBLUE   = "D5E8F0"
C_GREEN   = "D5EAD0"
C_YELLOW  = "FFFDD0"
C_RED     = "FFE0E0"
C_WHITE   = "FFFFFF"
C_ORANGE  = "FF8C00"

GRADE_COLORS = {
    "★★★★★": C_GREEN,
    "★★★★":  C_MBLUE,
    "★★★":   C_YELLOW,
    "★★":    "FFE5CC",
    "★":     C_RED,
    "—":     C_LGRAY,
}

PLATFORM_COLORS = {
    "TikTok":    "E8F4FD",
    "Instagram": "FDE8F4",
    "YouTube":   "FDF0E8",
    "Twitter":   "E8EFFE",
    "Lipscosme": "F4E8FD",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="222222", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)


def _border(style="thin", color="CCCCCC"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ──────────────────────────────────────────────────────────────
#  Sheet 1: KOL 스코어카드
# ──────────────────────────────────────────────────────────────

SCORECARD_HEADERS = [
    ("KOL명",        14),
    ("플랫폼",       10),
    ("비용(JPY)",    12),
    ("분석\n게시물수", 8),
    ("평균\n조회수",  10),
    ("평균\n좋아요",  10),
    ("평균\n댓글",    9),
    ("평균\n저장",    9),
    ("평균\n공유",    9),
    ("CPV\n(¥/회)",  10),
    ("ER%",          8),
    ("저장률%",       9),
    ("CPE\n(¥/건)",  10),
    ("저장비율%",     9),
    ("댓글비율%",     9),
    ("종합점수",      10),
    ("등급",          10),
    ("채택권고",      14),
]

NUM_FMT_COMMA = '#,##0'
NUM_FMT_DEC1  = '#,##0.0'
NUM_FMT_DEC2  = '#,##0.00'
NUM_FMT_PCT   = '0.00%'


def _write_scorecard_sheet(ws, kols_df):
    """Sheet1: 스코어카드"""
    # 제목 행
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = "KOL 선별 스코어카드"
    title_cell.font = _font(bold=True, size=14, color=C_WHITE)
    title_cell.fill = _fill(C_NAVY)
    title_cell.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # 헤더 행
    header_fill  = _fill(C_BLUE)
    header_font  = _font(bold=True, size=9, color=C_WHITE)
    header_align = _align("center", wrap=True)
    ws.row_dimensions[2].height = 32

    for col_idx, (header, _) in enumerate(SCORECARD_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.border = _border("thin", "4472C4")
        cell.alignment = header_align

    # 컬럼 너비
    for col_idx, (_, width) in enumerate(SCORECARD_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 행
    COL_NAMES = [
        "KOL명", "플랫폼", "비용(JPY)", "분석게시물수",
        "평균 조회수", "평균 좋아요", "평균 댓글", "평균 저장", "평균 공유",
        "CPV(¥/회)", "ER%", "저장률%", "CPE(¥/건)", "저장비율%", "댓글비율%",
        "종합점수", "등급", "채택권고"
    ]
    NUMBER_COLS = {
        3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16
    }  # 1-indexed

    for row_idx, row in kols_df.iterrows():
        excel_row = row_idx + 3  # row 3부터 시작
        grade_val = str(row.get("등급", "—"))
        plat_val  = str(row.get("플랫폼", ""))

        row_bg = GRADE_COLORS.get(grade_val, C_WHITE)
        plat_bg = PLATFORM_COLORS.get(plat_val, C_WHITE)

        for col_idx, col_name in enumerate(COL_NAMES, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            val  = row.get(col_name) if col_name in row.index else None
            cell.border = _border()

            # 배경색 (플랫폼 계열 + 등급 강조 혼합)
            if col_idx in (17, 18):  # 등급, 채택권고
                cell.fill = _fill(row_bg)
                cell.font = _font(bold=True, size=9)
            elif col_idx == 2:       # 플랫폼
                cell.fill = _fill(plat_bg)
                cell.font = _font(bold=False, size=9)
            else:
                bg = C_LGRAY if row_idx % 2 == 0 else C_WHITE
                cell.fill = _fill(bg)
                cell.font = _font(size=9)

            # 값 설정 및 서식
            if val is None or (isinstance(val, float) and str(val) == "nan"):
                cell.value = "—"
                cell.alignment = _align("center")
            elif col_idx == 3:  # 비용
                cell.value = float(val)
                cell.number_format = '#,##0"¥"'
                cell.alignment = _align("right")
            elif col_idx in (5, 6, 7, 8, 9):  # 평균 지표 (정수)
                cell.value = int(round(float(val))) if val else "—"
                cell.number_format = NUM_FMT_COMMA
                cell.alignment = _align("right")
            elif col_idx in (10, 13):  # CPV, CPE
                cell.value = float(val)
                cell.number_format = NUM_FMT_DEC1
                cell.alignment = _align("right")
            elif col_idx in (11, 12, 14, 15):  # ER%, 저장률% 등
                cell.value = float(val)
                cell.number_format = NUM_FMT_DEC2
                cell.alignment = _align("right")
            elif col_idx == 16:  # 종합점수
                cell.value = float(val)
                cell.number_format = NUM_FMT_DEC1
                cell.alignment = _align("center")
                cell.font = _font(bold=True, size=10)
            elif col_idx == 4:  # 게시물 수
                cell.value = int(val)
                cell.alignment = _align("center")
            else:
                cell.value = val
                cell.alignment = _align("left") if col_idx == 1 else _align("center")

        ws.row_dimensions[excel_row].height = 18

    # 열 고정 (1행 헤더 + 이름 열 고정)
    ws.freeze_panes = "C3"


# ──────────────────────────────────────────────────────────────
#  Sheet 2: 플랫폼별 랭킹
# ──────────────────────────────────────────────────────────────

def _write_ranking_sheet(ws, kols_df):
    """Sheet2: 플랫폼별 랭킹"""
    ws.merge_cells("A1:F1")
    ws["A1"].value = "플랫폼별 종합 랭킹"
    ws["A1"].font  = _font(bold=True, size=13, color=C_WHITE)
    ws["A1"].fill  = _fill(C_TEAL)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 26

    current_row = 3
    platforms = kols_df["플랫폼"].dropna().unique().tolist()
    ordered = [p for p in ["TikTok","Instagram","YouTube","Twitter","Lipscosme"]
               if p in platforms]
    ordered += [p for p in platforms if p not in ordered]

    for plat in ordered:
        plat_df = (kols_df[kols_df["플랫폼"] == plat]
                   .copy()
                   .dropna(subset=["종합점수"])
                   .sort_values("종합점수", ascending=False)
                   .reset_index(drop=True))

        if plat_df.empty:
            continue

        # 플랫폼 헤더
        ws.merge_cells(f"A{current_row}:F{current_row}")
        c = ws.cell(row=current_row, column=1, value=f"  {plat}")
        c.font  = _font(bold=True, size=11, color=C_WHITE)
        c.fill  = _fill(C_NAVY)
        c.alignment = _align("left")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # 서브 헤더
        for ci, hdr in enumerate(["순위", "KOL명", "종합점수", "등급", "채택권고", "주요지표"], 1):
            cell = ws.cell(row=current_row, column=ci, value=hdr)
            cell.font   = _font(bold=True, size=9, color=C_WHITE)
            cell.fill   = _fill(C_BLUE)
            cell.border = _border()
            cell.alignment = _align("center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # 데이터
        for rank_idx, (_, row) in enumerate(plat_df.iterrows(), start=1):
            grade = str(row.get("등급", "—"))
            bg = GRADE_COLORS.get(grade, C_WHITE)

            rank_cell = ws.cell(row=current_row, column=1, value=f"#{rank_idx}")
            rank_cell.alignment = _align("center")
            rank_cell.font   = _font(bold=True, size=10)
            rank_cell.fill   = _fill(bg)
            rank_cell.border = _border()

            name_cell = ws.cell(row=current_row, column=2,
                                value=row.get("KOL명", ""))
            name_cell.alignment = _align("left")
            name_cell.font   = _font(bold=(rank_idx == 1), size=10)
            name_cell.fill   = _fill(bg)
            name_cell.border = _border()

            score_val = row.get("종합점수")
            score_cell = ws.cell(row=current_row, column=3,
                                 value=float(score_val) if score_val else "—")
            score_cell.number_format = "0.0"
            score_cell.alignment = _align("center")
            score_cell.font   = _font(bold=True, size=11)
            score_cell.fill   = _fill(bg)
            score_cell.border = _border()

            grade_cell = ws.cell(row=current_row, column=4, value=grade)
            grade_cell.alignment = _align("center")
            grade_cell.font   = _font(size=10)
            grade_cell.fill   = _fill(bg)
            grade_cell.border = _border()

            adopt_cell = ws.cell(row=current_row, column=5,
                                 value=row.get("채택권고", ""))
            adopt_cell.alignment = _align("center")
            adopt_cell.font   = _font(size=9)
            adopt_cell.fill   = _fill(bg)
            adopt_cell.border = _border()

            # 주요 지표 요약
            metric_parts = []
            if plat == "TikTok":
                cpv = row.get("CPV(¥/회)")
                er  = row.get("ER%")
                sr  = row.get("저장률%")
                if cpv: metric_parts.append(f"CPV ¥{cpv:.1f}")
                if er:  metric_parts.append(f"ER {er:.1f}%")
                if sr:  metric_parts.append(f"저장 {sr:.2f}%")
            elif plat == "Instagram":
                cpe = row.get("CPE(¥/건)")
                cr  = row.get("댓글비율%")
                if cpe: metric_parts.append(f"CPE ¥{cpe:.0f}")
                if cr:  metric_parts.append(f"댓글비율 {cr:.1f}%")
            elif plat == "YouTube":
                cpv = row.get("CPV(¥/회)")
                if cpv: metric_parts.append(f"CPV ¥{cpv:.1f}")
            elif plat == "Twitter":
                cpv = row.get("CPV(¥/회)")
                er  = row.get("ER%")
                if cpv: metric_parts.append(f"CPV ¥{cpv:.1f}")
                if er:  metric_parts.append(f"ER {er:.1f}%")
            elif plat == "Lipscosme":
                cpe = row.get("CPE(¥/건)")
                sv  = row.get("저장비율%")
                if cpe: metric_parts.append(f"CPE ¥{cpe:.0f}")
                if sv:  metric_parts.append(f"저장비율 {sv:.1f}%")

            metric_cell = ws.cell(row=current_row, column=6,
                                  value=" / ".join(metric_parts) if metric_parts else "—")
            metric_cell.alignment = _align("left")
            metric_cell.font   = _font(size=9)
            metric_cell.fill   = _fill(bg)
            metric_cell.border = _border()

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 2  # 플랫폼 간 간격

    # 열 너비
    _set_col_widths(ws, {
        "A": 7, "B": 18, "C": 10, "D": 12, "E": 16, "F": 35
    })


# ──────────────────────────────────────────────────────────────
#  Sheet 3: 벤치마크 기준
# ──────────────────────────────────────────────────────────────

BENCHMARK_DATA = [
    ("TikTok",     "평균 ER%",    "5~9%"),
    ("TikTok",     "우수 ER%",    "10%+"),
    ("TikTok",     "평균 CPV",    "¥1~3"),
    ("TikTok",     "저장률(뷰티)","1~3%"),
    ("Instagram",  "피드 ER%",    "3~5%"),
    ("Instagram",  "릴스 ER%",    "4~8%"),
    ("Instagram",  "우수 ER%",    "12%+"),
    ("Instagram",  "저장률(뷰티)","1~3%"),
    ("YouTube",    "숏츠 CPV",    "¥0.3~1.5"),
    ("YouTube",    "롱폼 CPV",    "¥2~8"),
    ("YouTube",    "평균 ER%",    "2~4%"),
    ("Twitter",    "평균 ER%",    "0.5~2%"),
    ("Twitter",    "CPV",         "¥0.5~2"),
    ("Lipscosme",  "저장비율",    "15~30%"),
    ("Lipscosme",  "게시물당 좋아요","50~200"),
    ("Lipscosme",  "게시물당 저장","10~50"),
]

WEIGHT_DATA = [
    ("TikTok",    "CPV 35% + ER% 35% + 저장률 30%"),
    ("Instagram", "CPE 60% + 댓글비율 40%"),
    ("YouTube",   "CPV 100%"),
    ("Twitter",   "CPV 50% + ER% 30% + CPE 20%"),
    ("Lipscosme", "CPE 50% + 저장비율 50%"),
]


def _write_benchmark_sheet(ws):
    """Sheet3: 벤치마크 기준"""
    # ── 벤치마크 테이블 ──
    ws["A1"].value = "일본 뷰티 카테고리 벤치마크 (2024-2025)"
    ws["A1"].font  = _font(bold=True, size=12, color=C_WHITE)
    ws["A1"].fill  = _fill(C_NAVY)
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 22

    for ci, hdr in enumerate(["플랫폼", "지표", "기준값"], 1):
        cell = ws.cell(row=2, column=ci, value=hdr)
        cell.font   = _font(bold=True, size=9, color=C_WHITE)
        cell.fill   = _fill(C_BLUE)
        cell.border = _border()
        cell.alignment = _align("center")

    for ri, (plat, metric, val) in enumerate(BENCHMARK_DATA, start=3):
        bg = PLATFORM_COLORS.get(plat, C_WHITE)
        for ci, v in enumerate([plat, metric, val], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill   = _fill(bg)
            cell.border = _border()
            cell.font   = _font(size=9)
            cell.alignment = _align("center" if ci != 2 else "left")
        ws.row_dimensions[ri].height = 16

    # ── 가중치 테이블 ──
    start_row = len(BENCHMARK_DATA) + 4
    ws.merge_cells(f"A{start_row}:B{start_row}")
    ws.cell(row=start_row, column=1, value="플랫폼별 가중치 요약").font = \
        _font(bold=True, size=11, color=C_WHITE)
    ws.cell(row=start_row, column=1).fill = _fill(C_TEAL)
    ws.cell(row=start_row, column=1).alignment = _align("center")
    ws.row_dimensions[start_row].height = 20

    for ci, hdr in enumerate(["플랫폼", "가중치 (낮은 CPV 또는 높은 지표 = 우수)"], 1):
        cell = ws.cell(row=start_row + 1, column=ci, value=hdr)
        cell.font   = _font(bold=True, size=9, color=C_WHITE)
        cell.fill   = _fill(C_BLUE)
        cell.border = _border()
        cell.alignment = _align("center")

    for ri2, (plat, weights) in enumerate(WEIGHT_DATA, start=start_row + 2):
        bg = PLATFORM_COLORS.get(plat, C_WHITE)
        for ci, v in enumerate([plat, weights], 1):
            cell = ws.cell(row=ri2, column=ci, value=v)
            cell.fill   = _fill(bg)
            cell.border = _border()
            cell.font   = _font(size=9)
            cell.alignment = _align("center" if ci == 1 else "left")
        ws.row_dimensions[ri2].height = 16

    # ── 등급 설명 ──
    grade_start = start_row + len(WEIGHT_DATA) + 3
    ws.merge_cells(f"A{grade_start}:C{grade_start}")
    ws.cell(row=grade_start, column=1, value="5단계 등급 기준").font = \
        _font(bold=True, size=11, color=C_WHITE)
    ws.cell(row=grade_start, column=1).fill = _fill(C_NAVY)
    ws.cell(row=grade_start, column=1).alignment = _align("center")

    grades = [
        ("★★★★★", "8.5~10.0", "최우선 채택 - 동일 플랫폼 탑티어"),
        ("★★★★",  "7.0~8.4",  "채택 권고 - 평균 상위"),
        ("★★★",   "5.5~6.9",  "조건부 채택 - 단가 협상 후"),
        ("★★",    "3.0~5.4",  "보류 - 대안 모색"),
        ("★",     "0~2.9",    "비권고 - 효율 현저히 낮음"),
    ]
    grade_colors = [C_GREEN, C_MBLUE, C_YELLOW, "FFE5CC", C_RED]
    for gi, ((grade, score_range, desc), bg) in enumerate(
        zip(grades, grade_colors), start=grade_start + 1
    ):
        ws.cell(row=gi, column=1, value=grade).fill   = _fill(bg)
        ws.cell(row=gi, column=1).font   = _font(bold=True, size=10)
        ws.cell(row=gi, column=1).border = _border()
        ws.cell(row=gi, column=1).alignment = _align("center")
        ws.cell(row=gi, column=2, value=score_range).fill   = _fill(bg)
        ws.cell(row=gi, column=2).border = _border()
        ws.cell(row=gi, column=2).alignment = _align("center")
        ws.cell(row=gi, column=3, value=desc).fill   = _fill(bg)
        ws.cell(row=gi, column=3).border = _border()
        ws.cell(row=gi, column=3).alignment = _align("left")
        ws.row_dimensions[gi].height = 16

    _set_col_widths(ws, {"A": 12, "B": 38, "C": 35})


# ──────────────────────────────────────────────────────────────
#  메인 내보내기 함수
# ──────────────────────────────────────────────────────────────

def generate_excel_scorecard(kols_df) -> bytes:
    """
    KOL DataFrame → Excel 바이트 스트림 반환
    (Streamlit st.download_button 에 직접 전달 가능)
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # Sheet 1: 스코어카드
    ws1 = wb.create_sheet("KOL 스코어카드")
    _write_scorecard_sheet(ws1, kols_df)

    # Sheet 2: 플랫폼별 랭킹
    ws2 = wb.create_sheet("플랫폼별 랭킹")
    _write_ranking_sheet(ws2, kols_df)

    # Sheet 3: 벤치마크 기준
    ws3 = wb.create_sheet("벤치마크 기준")
    _write_benchmark_sheet(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
